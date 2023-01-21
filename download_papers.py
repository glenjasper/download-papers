#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import re
import sys
import time
import shutil
import argparse
import traceback
import xlsxwriter
import subprocess
from openpyxl import load_workbook

def menu(args):
    parser = argparse.ArgumentParser(description = "This scripts downloads .pdf files from formatted .xlsx files, via DOI.", epilog = "Thank you!")
    parser.add_argument("-i", "--input_file", required = True, help = ".xlsx file that contains the DOIs")
    parser.add_argument("-o", "--output", help = "Output folder")
    parser.add_argument("--version", action = "version", version = "%s %s" % ('%(prog)s', oscihub.VERSION))
    args = parser.parse_args()

    # Check scidownl
    out_scidownl = oscihub.get_command('scidownl -h')
    if 'Usage: scidownl' not in out_scidownl:
        oscihub.show_print("It looks like 'scidownl' is not installed, you can install it with: pip3 install -U scidownl", showdate = False, font = oscihub.YELLOW)
        exit()

    file_name = os.path.basename(args.input_file)
    file_path = os.path.dirname(args.input_file)
    if file_path is None or file_path == "":
        file_path = os.getcwd().strip()

    oscihub.INPUT_FILE = os.path.join(file_path, file_name)
    if not oscihub.check_path(oscihub.INPUT_FILE):
        oscihub.show_print("%s: error: the file '%s' doesn't exist" % (os.path.basename(__file__), oscihub.INPUT_FILE), showdate = False, font = oscihub.YELLOW)
        oscihub.show_print("%s: error: the following arguments are required: -i/--input_file" % os.path.basename(__file__), showdate = False, font = oscihub.YELLOW)
        exit()

    if args.output is not None:
        output_name = os.path.basename(args.output)
        output_path = os.path.dirname(args.output)
        if output_path is None or output_path == "":
            output_path = os.getcwd().strip()

        oscihub.OUTPUT_PATH = os.path.join(output_path, output_name)
        created = oscihub.create_directory(oscihub.OUTPUT_PATH)
        if not created:
            oscihub.show_print("%s: error: Couldn't create folder '%s'" % (os.path.basename(__file__), oscihub.OUTPUT_PATH), showdate = False, font = oscihub.YELLOW)
            exit()
    else:
        oscihub.OUTPUT_PATH = os.getcwd().strip()
        oscihub.OUTPUT_PATH = os.path.join(oscihub.OUTPUT_PATH, 'output_download')
        oscihub.create_directory(oscihub.OUTPUT_PATH)

class SCIhub:

    def __init__(self):
        self.VERSION = 1.0

        self.INPUT_FILE = None
        self.OUTPUT_PATH = None

        self.ROOT_DIR = os.path.dirname(os.path.realpath(__file__))
        self.LOG_NAME = "run_%s_%s.log" % (os.path.splitext(os.path.basename(__file__))[0], time.strftime('%Y%m%d'))
        self.LOG_FILE = None
        self.SUMMARY_FILE_CONTROL = 'summary_control.txt'

        # SciHub
        self.SCIHUB_URL = 'https://sci-hub.tw'
        self.SCIHUB_ID = 0 # 2

        # Folder
        self.FOLDER_TXT = 'Papers'
        self.FOLDER_TEMP = 'temporal_folder'

        # Year
        self.STATUS_NO_YEAR = 'NoYear'

        # Status
        self.STATUS_OK = 'Ok'
        self.STATUS_NOT_AVAILABLE = 'Not available'
        self.STATUS_NONEXISTENT = 'Non-existent'
        self.STATUS_NAME = 'Status'

        # Types
        self.TYPE_INPUT = None
        self.TYPE_REPOSITORY_UNION = "repository_union"
        self.TYPE_REPOSITORY_UNIQUE = "repository_unique"
        self.TYPE_TXT = "txt"

        # Xls Summary
        self.XLS_FILE = 'summary_download.xlsx'
        self.XLS_SHEET_DETAIL = 'Detail'

        # Xls Columns
        self.xls_col_item = 'Item'
        self.xls_col_title = 'Title'
        self.xls_col_year = 'Year'
        self.xls_col_doi = 'DOI'
        self.xls_col_document_type = 'Document Type'
        self.xls_col_languaje = 'Language'
        self.xls_col_cited_by = 'Cited By'
        self.xls_col_download = 'Download'
        self.xls_col_authors = 'Author(s)'
        self.xls_col_repository = 'Repository'
        self.xls_col_pdf_name = 'PDF Name'

        self.xls_columns_csv = [self.xls_col_item,
                                self.xls_col_title,
                                self.xls_col_year,
                                self.xls_col_doi,
                                self.xls_col_document_type,
                                self.xls_col_languaje,
                                self.xls_col_cited_by,
                                self.xls_col_download,
                                self.xls_col_authors]

        self.xls_columns_txt = [self.xls_col_item,
                                self.xls_col_doi,
                                self.xls_col_download]

        self.default_document_type = 'Unknown Type'

        # Fonts
        self.RED = '\033[31m'
        self.GREEN = '\033[32m'
        self.YELLOW = '\033[33m'
        self.BIRED = '\033[1;91m'
        self.BIGREEN = '\033[1;92m'
        self.END = '\033[0m'

    def show_print(self, message, logs = None, showdate = True, font = None):
        msg_print = message
        msg_write = message

        if font is not None:
            msg_print = "%s%s%s" % (font, msg_print, self.END)

        if showdate is True:
            _time = time.strftime('%Y-%m-%d %H:%M:%S')
            msg_print = "%s %s" % (_time, msg_print)
            msg_write = "%s %s" % (_time, message)

        print(msg_print)
        if logs is not None:
            for log in logs:
                if log is not None:
                    with open(log, 'a', encoding = 'utf-8') as f:
                        f.write("%s\n" % msg_write)
                        f.close()

    def start_time(self):
        return time.time()

    def finish_time(self, start, message = None):
        finish = time.time()
        runtime = time.strftime("%H:%M:%S", time.gmtime(finish - start))
        if message is None:
            return runtime
        else:
            return "%s: %s" % (message, runtime)

    def create_directory(self, path):
        output = True
        try:
            if len(path) > 0 and not os.path.exists(path):
                os.makedirs(path)
        except Exception as e:
            output = False
        return output

    def check_path(self, path):
        _check = False
        if path is not None:
            if len(path) > 0 and os.path.exists(path):
                _check = True
        return _check

    def check_title(self, title):
        rstr = r"[\/\\\:\*\?\"\“\”\<\>\|\@\°\'\‘\’\®]" # / \ : * ? " “ ” < > | @ ° ' ‘ ’ ®
        new_title = re.sub(rstr, " ", title)[:200]
        new_title = re.sub("\n", " ", new_title)
        new_title = re.sub("  ", " ", new_title)

        return new_title

    def remove_directory(self, path):
        output = True
        try:
            if self.check_path(path):
                shutil.rmtree(path)
        except Exception as e:
            output = False
        return output

    def set_xls_type(self):
        workbook = load_workbook(filename = self.INPUT_FILE, data_only = True)
        sheet = workbook[self.XLS_SHEET_DETAIL]
        rows = sheet.rows

        ncolumns = 0
        for row in rows:
            ncolumns = len(row)
            break

        if ncolumns == 9:
            self.TYPE_INPUT = self.TYPE_REPOSITORY_UNION
        elif ncolumns == 8:
            self.TYPE_INPUT = self.TYPE_REPOSITORY_UNIQUE
        elif ncolumns == 2:
            self.TYPE_INPUT = self.TYPE_TXT

    def read_xls_summary(self):
        workbook = load_workbook(filename = self.INPUT_FILE, data_only = True)
        sheet = workbook[self.XLS_SHEET_DETAIL]
        rows = sheet.rows

        file_collection = {}
        for index_i, row in enumerate(rows):
            collection = {}
            for index_j, cell in enumerate(row):
                if cell.value == self.xls_col_item:
                    break
                column_name = None

                if self.TYPE_INPUT == self.TYPE_TXT:
                    if index_j == 0:
                        column_name = self.xls_col_item
                    elif index_j == 1:
                        column_name = self.xls_col_doi
                else:
                    if index_j == 0:
                        column_name = self.xls_col_item
                    elif index_j == 1:
                        column_name = self.xls_col_title
                    elif index_j == 2:
                        column_name = self.xls_col_year
                    elif index_j == 3:
                        column_name = self.xls_col_doi
                    elif index_j == 4:
                        column_name = self.xls_col_document_type
                    elif index_j == 5:
                        column_name = self.xls_col_languaje
                    elif index_j == 6:
                        column_name = self.xls_col_cited_by
                    elif index_j == 7:
                        column_name = self.xls_col_authors
                    elif index_j == 8:
                        column_name = self.xls_col_repository

                collection.update({column_name: cell.value})

            if len(collection) > 0:
                file_collection.update({index_i: collection})

        return file_collection

    def write_file_control(self, doi, status):
        if self.TYPE_INPUT == self.TYPE_TXT:
            open(self.SUMMARY_FILE_CONTROL, 'a').write('%s\t%s\n' % (doi, status))

    def save_summary_xls(self, data_paper, data_status):
        if self.TYPE_INPUT == self.TYPE_TXT:
            _xls_columns = self.xls_columns_txt.copy()
        else:
            if self.TYPE_INPUT == self.TYPE_REPOSITORY_UNION:
                self.xls_columns_csv.append(self.xls_col_repository)
            self.xls_columns_csv.append(self.xls_col_pdf_name)

            _xls_columns = self.xls_columns_csv.copy()

        _last_col = len(_xls_columns) - 1

        workbook = xlsxwriter.Workbook(self.XLS_FILE)
        worksheet = workbook.add_worksheet(self.XLS_SHEET_DETAIL)
        worksheet.freeze_panes(row = 1, col = 0) # Freeze the first row.
        worksheet.autofilter(first_row = 0, first_col = 0, last_row = 0, last_col = _last_col) # 'A1:H1'
        worksheet.set_default_row(height = 14.5)

        # Add columns
        cell_format_title = workbook.add_format({'bold': True,
                                                 'font_color': 'white',
                                                 'bg_color': 'black',
                                                 'align': 'center',
                                                 'valign': 'vcenter'})
        for icol, column in enumerate(_xls_columns):
            worksheet.write(0, icol, column, cell_format_title)

        # Add rows
        if self.TYPE_INPUT == self.TYPE_TXT:
            worksheet.set_column(first_col = 0, last_col = 0, width = 7)  # Column A:A
            worksheet.set_column(first_col = 1, last_col = 1, width = 33) # Column B:B
            worksheet.set_column(first_col = 2, last_col = 2, width = 13) # Column C:C
        else:
            worksheet.set_column(first_col = 0, last_col = 0, width = 7)  # Column A:A
            worksheet.set_column(first_col = 1, last_col = 1, width = 40) # Column B:B
            worksheet.set_column(first_col = 2, last_col = 2, width = 8)  # Column C:C
            worksheet.set_column(first_col = 3, last_col = 3, width = 33) # Column D:D
            worksheet.set_column(first_col = 4, last_col = 4, width = 18) # Column E:E
            worksheet.set_column(first_col = 5, last_col = 5, width = 12) # Column F:F
            worksheet.set_column(first_col = 6, last_col = 6, width = 11) # Column G:G
            worksheet.set_column(first_col = 7, last_col = 7, width = 13) # Column H:H
            worksheet.set_column(first_col = 8, last_col = 8, width = 36) # Column I:I
            if self.TYPE_INPUT == self.TYPE_REPOSITORY_UNION:
                worksheet.set_column(first_col = 9, last_col = 9, width = 13) # Column J:J
                worksheet.set_column(first_col = 10, last_col = 10, width = 30) # Column K:K
            else:
                worksheet.set_column(first_col = 9, last_col = 9, width = 30) # Column J:J

        cell_format_row = workbook.add_format({'text_wrap': True, 'valign': 'top'})
        icol = 0
        for irow, item in data_paper.items():
            if self.TYPE_INPUT == self.TYPE_TXT:
                col_doi = item[self.xls_col_doi]
                ctrl_title = col_doi
            else:
                col_title = item[self.xls_col_title]
                col_year = item[self.xls_col_year]
                col_doi = item[self.xls_col_doi]
                col_document_type = item[self.xls_col_document_type]
                col_document_type = self.default_document_type if col_document_type is None else col_document_type
                col_languaje = item[self.xls_col_languaje]
                col_cited_by = item[self.xls_col_cited_by]
                col_authors = item[self.xls_col_authors]
                if self.TYPE_INPUT == self.TYPE_REPOSITORY_UNION:
                    col_repository = item[self.xls_col_repository]

                _col_year = self.STATUS_NO_YEAR if col_year is None else col_year
                _year_title = '%s.%s' % (_col_year, col_title)
                ctrl_title = '%s.%s.pdf' % (col_document_type, self.check_title(_year_title))

                col_pdf_name = None
                if data_status[ctrl_title] == self.STATUS_OK:
                    col_pdf_name = item[self.xls_col_pdf_name]

            if self.TYPE_INPUT == self.TYPE_TXT:
                worksheet.write(irow, icol + 0, irow, cell_format_row)
                worksheet.write(irow, icol + 1, col_doi, cell_format_row)
                worksheet.write(irow, icol + 2, data_status[ctrl_title], cell_format_row)
            else:
                worksheet.write(irow, icol + 0, irow, cell_format_row)
                worksheet.write(irow, icol + 1, col_title, cell_format_row)
                worksheet.write(irow, icol + 2, col_year, cell_format_row)
                worksheet.write(irow, icol + 3, col_doi, cell_format_row)
                worksheet.write(irow, icol + 4, col_document_type, cell_format_row)
                worksheet.write(irow, icol + 5, col_languaje, cell_format_row)
                worksheet.write(irow, icol + 6, col_cited_by, cell_format_row)
                worksheet.write(irow, icol + 7, data_status[ctrl_title], cell_format_row)
                worksheet.write(irow, icol + 8, col_authors, cell_format_row)
                if self.TYPE_INPUT == self.TYPE_REPOSITORY_UNION:
                    worksheet.write(irow, icol + 9, col_repository, cell_format_row)
                    worksheet.write(irow, icol + 10, col_pdf_name, cell_format_row)
                else:
                    worksheet.write(irow, icol + 9, col_pdf_name, cell_format_row)
        workbook.close()

    def save_summary_text(self, dict_control):
        if self.TYPE_INPUT == self.TYPE_TXT:
            self.remove_file(self.SUMMARY_FILE_CONTROL)
            if len(dict_control) > 0:
                self.save_text_file(dictionary = dict_control, file = self.SUMMARY_FILE_CONTROL, title = 'Status')

    def remove_file(self, file):
        if self.check_path(file):
            os.remove(file)

    def save_text_file(self, dictionary, file, title):
        with open(file, 'w') as f:
            f.write('Index\t%s\n' % title)
            for i, value in dictionary.items():
                f.write('%s\t%s\n' % (i, value))
        f.close()

    def get_expected_files(self, dictionary):
        folders = {}
        for idx, item in dictionary.items():
            _document_type = item[self.xls_col_document_type]
            _document_type = self.default_document_type if _document_type is None else _document_type
            _year = item[self.xls_col_year]
            _year = self.STATUS_NO_YEAR if _year is None else _year 
            _title = item[self.xls_col_title]
            _year_title = '%s.%s' % (_year, _title)
            _title = self.check_title(_year_title)
            pdfname = '%s.pdf' % (_title)

            item.update({self.xls_col_pdf_name: pdfname})

            if _document_type not in folders:
                documents = [pdfname]
            else:
                documents = folders[_document_type].copy()
                documents.append(pdfname)

            folders.update({_document_type: documents})

        return folders

    def get_downloaded_files(self, dictionary = None):
        summary_ctrl = {}
        if self.TYPE_INPUT == self.TYPE_TXT:
            if self.check_path(self.SUMMARY_FILE_CONTROL):
                with open(self.SUMMARY_FILE_CONTROL, 'r') as f:
                    for line in f:
                        line = line.strip()
                        if not line.startswith('Index'):
                            line = line.split('\t')
                            summary_ctrl.update({line[0]: line[1]})
                f.close()
            else:
                open(self.SUMMARY_FILE_CONTROL, 'w').close()
        else:
            self.remove_directory(self.FOLDER_TEMP)
            self.create_directory(self.FOLDER_TEMP)

            for folder, files in dictionary.items():
                this_folder = os.path.join(self.OUTPUT_PATH, folder)
                if self.check_path(this_folder):
                    shutil.move(this_folder, self.FOLDER_TEMP)

            for folder, files in dictionary.items():
                for file in files:
                    this_file = os.path.join(self.FOLDER_TEMP, folder, file)
                    if self.check_path(this_file):
                        real_folder = os.path.join(self.OUTPUT_PATH, folder)
                        self.create_directory(real_folder)
                        shutil.move(this_file, real_folder)
                        ctrl_name = '%s.%s' % (folder, file)
                        summary_ctrl.update({ctrl_name: self.STATUS_OK})

            self.remove_directory(self.FOLDER_TEMP)

        return summary_ctrl

    def update_status(self, dictionary, dict_ctrl):
        for _, item in dictionary.items():
            if self.TYPE_INPUT == self.TYPE_TXT:
                _doi = item[self.xls_col_doi]
                _status = dict_ctrl[_doi] if _doi in dict_ctrl else None
            else:
                _document_type = item[self.xls_col_document_type]
                _document_type = self.default_document_type if _document_type is None else _document_type
                _year = item[self.xls_col_year]
                _year = self.STATUS_NO_YEAR if _year is None else _year 
                _title = item[self.xls_col_title]
                _year_title = '%s.%s' % (_year, _title)
                _title = self.check_title(_year_title)
                pdfname = '%s.%s.pdf' % (_document_type, _title)
                _status = dict_ctrl[pdfname] if pdfname in dict_ctrl else None
            item.update({self.STATUS_NAME: _status})

    def download_pdf(self, dict_information, dict_ctrl):
        record_count = len(dict_information)
        summary_not_availables = {}
        summary_non_existents = {}
        for idx, item in dict_information.items():
            doi = item[self.xls_col_doi]
            status = item[self.STATUS_NAME]

            if self.TYPE_INPUT == self.TYPE_TXT:
                title = None
                document_type = self.FOLDER_TXT
                message = "[%s/%s] Analyzing the DOI: %s" % (idx, record_count, doi)
                year = idx
                ctrl_title = doi
            else:
                title = item[self.xls_col_title]
                document_type = item[self.xls_col_document_type]
                document_type = self.default_document_type if document_type is None else document_type
                message = "[%s/%s] Analyzing the Paper: %s" % (idx, record_count, title)
                year = item[self.xls_col_year]
                year = self.STATUS_NO_YEAR if year is None else year
                _year_title = '%s.%s' % (year, title)
                ctrl_title = '%s.%s.pdf' % (document_type, self.check_title(_year_title))

            self.show_print(message, [self.LOG_FILE], font = self.YELLOW)

            if status == self.STATUS_OK:
                self.show_print("[%s/%s] Paper already downloaded" % (idx, record_count), [self.LOG_FILE], font = self.GREEN)
                self.show_print("", [self.LOG_FILE])
                continue
            elif status == self.STATUS_NONEXISTENT:
                self.show_print("[%s/%s] Paper without DOI" % (idx, record_count), [self.LOG_FILE], font = self.GREEN)
                self.show_print("", [self.LOG_FILE])
                summary_non_existents.update({idx: title})
                continue
            else:
                # For Status: None and Not available
                if doi is not None:
                    try:
                        directory = os.path.join(self.OUTPUT_PATH, document_type)
                        self.create_directory(directory)

                        self.show_print("[%s/%s] Downloading paper..." % (idx, record_count), [self.LOG_FILE], font = self.GREEN)
                        self.run_scidownl(doi = doi, out = directory, filename = '%s.%s' % (year, title))
                        self.show_print("", [self.LOG_FILE])
                        dict_ctrl.update({ctrl_title: self.STATUS_OK})
                        self.write_file_control(ctrl_title, self.STATUS_OK)

                        # Rename
                        pdf_downloaded = os.path.join(directory, '%s.%s.pdf' % (year, title.replace(' ', '_')))
                        pdf_downloaded_rename = os.path.join(directory, '%s.%s.pdf' % (year, title.replace('_', ' ')))
                        os.rename(pdf_downloaded, pdf_downloaded_rename)
                    except Exception:
                        self.show_print("[%s/%s] Download link not available, please try after sometime" % (idx, record_count), [self.LOG_FILE], font = self.YELLOW)
                        self.show_print("[%s/%s] Also try prepending 'http://dx.doi.org/' to input" % (idx, record_count), [self.LOG_FILE], font = self.YELLOW)
                        self.show_print("", [self.LOG_FILE])
                        summary_not_availables.update({idx: doi})
                        if status is None:
                            dict_ctrl.update({ctrl_title: self.STATUS_NOT_AVAILABLE})
                            self.write_file_control(ctrl_title, self.STATUS_NOT_AVAILABLE)
                else:
                    self.show_print("[%s/%s] Paper without DOI" % (idx, record_count), [self.LOG_FILE], font = self.YELLOW)
                    self.show_print("", [self.LOG_FILE])
                    summary_non_existents.update({idx: title})
                    dict_ctrl.update({ctrl_title: self.STATUS_NONEXISTENT})
                    self.write_file_control(ctrl_title, self.STATUS_NONEXISTENT)

        self.show_print("[SUMMARY]", [self.LOG_FILE], font = self.GREEN)
        self.show_print("  Papers/DOIs analyzed: %s" % record_count, [self.LOG_FILE], font = self.GREEN)
        self.show_print("    Papers/DOIs downloaded: %s (see %s)" % (record_count - len(summary_not_availables) - len(summary_non_existents), self.OUTPUT_PATH), [self.LOG_FILE], font = self.GREEN)
        self.save_summary_xls(dict_information, dict_ctrl)
        self.save_summary_text(dict_ctrl)
        self.show_print("  For more details see the file: %s" % self.XLS_FILE, [self.LOG_FILE], font = self.GREEN)

    def run_scidownl(self, doi, out, filename):
        # scidownl download --doi 10.1145/3375633 -o output_file

        command = ["scidownl download",
                   "--doi %s" % doi,
                   "--out %s" % os.path.join(out, filename.replace(' ', '_'))]

        # Command execution
        _command = " ".join(command)
        try:
            p = subprocess.Popen(_command, shell = True, stdout = subprocess.PIPE, stderr = subprocess.STDOUT)
        except Exception as e:
            self.show_print("Error %s while executing command %s" % (e, _command), [self.LOG_FILE], font = self.YELLOW)

        successful = False
        success_words = ["Successfully", "download"]
        for line in iter(p.stdout.readline, b''):
            _line = line.decode('utf-8').rstrip()
            if successful is False and self.search_word_array(success_words, _line):
                successful = True
            self.show_print(_line, [self.LOG_FILE])

        '''
        if not successful:
            self.show_print("ERROR executing!", [self.LOG_FILE], font = self.YELLOW)
            self.show_print("Check the command: %s" % (_command), [self.LOG_FILE], font = self.YELLOW)
            self.show_print("", [self.LOG_FILE])
            # sys.exit(1)
        '''
        assert successful, 'ERROR'

    def search_word_array(self, words = [], string = None):
        for item in words:
            if item not in string:
                return False
        return True

    def get_command(self, cmd):
        p = subprocess.Popen(cmd, shell = True, stdin = None, stdout = subprocess.PIPE, stderr = subprocess.PIPE)
        (checkStdout, checkStderr) = p.communicate()
        checkStdout = checkStdout.decode('utf-8').strip()
        checkStderr = checkStderr.decode('utf-8').strip()

        output_list = []
        if checkStdout:
            for line in checkStdout.splitlines():
                for word in line.strip().split():
                    output_list.append(word.strip())
        else:
            output_list.append('ERROR')

        return ' '.join(output_list)

def main(args):
    try:
        start = oscihub.start_time()
        menu(args)

        oscihub.LOG_FILE = os.path.join(oscihub.OUTPUT_PATH, oscihub.LOG_NAME)
        oscihub.XLS_FILE = os.path.join(oscihub.OUTPUT_PATH, oscihub.XLS_FILE)
        oscihub.FOLDER_TEMP = os.path.join(oscihub.OUTPUT_PATH, oscihub.FOLDER_TEMP)
        oscihub.SUMMARY_FILE_CONTROL = os.path.join(oscihub.OUTPUT_PATH, oscihub.SUMMARY_FILE_CONTROL)
        oscihub.set_xls_type()
        if oscihub.TYPE_INPUT is None:
            oscihub.show_print("The file is not in the correct format: %s" % oscihub.XLS_FILE, [oscihub.LOG_FILE])
            raise Exception('Incorrect format')

        oscihub.show_print("#############################################################################", [oscihub.LOG_FILE], font = oscihub.BIGREEN)
        oscihub.show_print("############################## Download papers ##############################", [oscihub.LOG_FILE], font = oscihub.BIGREEN)
        oscihub.show_print("#############################################################################", [oscihub.LOG_FILE], font = oscihub.BIGREEN)

        oscihub.show_print("Reading the .xls file: %s" % oscihub.XLS_FILE, [oscihub.LOG_FILE], font = oscihub.GREEN)
        input_information = oscihub.read_xls_summary()
        oscihub.show_print("  Records found: %s" % len(input_information), [oscihub.LOG_FILE])
        oscihub.show_print("", [oscihub.LOG_FILE])

        if oscihub.TYPE_INPUT == oscihub.TYPE_TXT:
            summary_ctrl = oscihub.get_downloaded_files()
        else:
            pdf_by_folders = oscihub.get_expected_files(input_information)
            summary_ctrl = oscihub.get_downloaded_files(pdf_by_folders)

        oscihub.update_status(input_information, summary_ctrl)
        oscihub.download_pdf(input_information, summary_ctrl)

        oscihub.show_print("", [oscihub.LOG_FILE])
        oscihub.show_print(oscihub.finish_time(start, "Elapsed time"), [oscihub.LOG_FILE])
        oscihub.show_print("Done!", [oscihub.LOG_FILE])
    except Exception as e:
        oscihub.show_print("\n%s" % traceback.format_exc(), [oscihub.LOG_FILE], font = oscihub.RED)
        oscihub.show_print(oscihub.finish_time(start, "Elapsed time"), [oscihub.LOG_FILE])
        oscihub.show_print("Done!", [oscihub.LOG_FILE])

if __name__ == '__main__':
    oscihub = SCIhub()
    main(sys.argv)
